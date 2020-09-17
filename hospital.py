import os
import openpyxl as xl
from openpyxl.styles import Font
from datetime import date
import re

class Hospital:
    def __init__(self):
       self.boldFont = Font(bold=True)
       self.sheets = ['Auth','Patients','Doctors','Departments','Appointments']
       self.workbook = self.__create_db_if_not_exists()

    def __create_db_if_not_exists(self):
        '''creates and initializes the excel file if not exists.'''
        if not os.path.isfile('db.xlsx'):
            workbook = xl.Workbook()
            for i in range(len(self.sheets)):
                sheet = workbook.create_sheet(self.sheets[i],i)
                sheet['A1'].font = self.boldFont
                if i==0:
                    sheet['A1'].value = 'Rule'
                    sheet['B1'].value = 'Secret'
                    sheet['B1'].font = self.boldFont
                    sheet['A2'].value = 'admin'
                    sheet['B2'].value = 'admin123'
                    sheet['A3'].value = 'user'
                    sheet['B3'].value = 'user123'
                else:
                    for j in ['A','B','C','D']: 
                        sheet.column_dimensions[j].width=22
                        sheet[j+str(1)].font = self.boldFont
            workbook['Patients']['A1'].value = 'ID'
            workbook['Patients']['B1'].value = 'NAME'
            workbook['Patients']['C1'].value = 'ADDRESS'
            workbook['Patients']['D1'].value = 'AGE'

            workbook['Doctors']['A1'].value = 'ID'
            workbook['Doctors']['B1'].value = 'NAME'
            workbook['Doctors']['C1'].value = 'DEPARTMENT'

            workbook['Appointments']['A1'].value = 'ID'
            workbook['Appointments']['B1'].value = 'DOCTOR'
            workbook['Appointments']['C1'].value = 'PATIENT'
            workbook['Appointments']['D1'].value = 'DATE'

            workbook['Departments']['A1'].value = 'ID'
            workbook['Departments']['B1'].value = 'NAME'
            workbook.save('db.xlsx')
            return workbook
        else: return xl.load_workbook('db.xlsx')

    def __validate(self,field,pattern,err):
        '''Validates user input based on a pattern.'''
        while True:
            x = input('\nEnter '+field+': ')
            if pattern.match(str(x)): return x
            else: print(err)

    def __authenticate(self,level,pwd):
        '''Authenticates Admin/User by password.'''
        auth = self.workbook['Auth']
        return auth[level].value == pwd

    def __show_admin_menu(self):
        '''Shows management menu for the Admin.'''
        while True:
            if self.rule=='admin':
                x = input('\n1: Manage patients\n2: Manage doctors\n3: Manage appointments\n4: Manage departments\n0: To go back\n\n'+self.rule+'$ ')
                if x=='0': break
                elif x == '1':  self.__show_action_menu('Patients')
                elif x == '2':  self.__show_action_menu('Doctors')
                elif x == '3':  self.__show_action_menu('Appointments')
                elif x == '4':  self.__show_action_menu('Departments')
                else: print('\n*****Invalid input!*****\n')    
            else: 
                x = input('\n1: View patients\n2: View doctors\n3: View appointments\n4: View departments\n0: To go back\n\n'+self.rule+'$ ')
                if x=='0': break
                elif x=='1':
                    self.__view('Patients')
                elif x=='2':
                    self.__view('Doctors')
                elif x=='3':
                    self.__view('Appointments')
                elif x=='4':
                    self.__view('Departments')
                else: print('\n*****Invalid input!*****')    

    def __show_action_menu(self,title):
        '''Shows action menu in admin menu items.'''
        while True:
            x = input('\n1: Add '+title+'\n2: Delete '+title+'\n3: View all '+title+'\n0: To go back\n\n'+self.rule+'$ ')
            if x == '0': break
            elif x == '1': self.__prompt_addition(title)
            elif x == '2': 
                x = self.__validate(title+' id',re.compile(r'\d+'),'*****Invalid ID! id must be a number*****')
                self.__delete_record(title,int(x))
            elif x == '3': self.__view(title)
            else: print('\n*****Invalid input!*****\n')

    def __create_dependent_field_if_not_exists(self,title,data):
        '''Checks if some dependent field of a sheet exists in its own sheet.'''
        if title == 'Doctors':
            sheet = self.workbook['Departments']
            for i in range(1,sheet.max_row+1):
                if sheet.cell(row=i,column=2).value == data[1]: return True
            else:
                print('\n*****No department registerd with the name of '+data[1]+'*****')
                print('*****please add the department first.*****')
                return False
        else:
            sheet = self.workbook['Doctors']
            for i in range(1,sheet.max_row+1):
                if sheet.cell(row=i,column=2).value == data[0]: return True
            else:
                print('\n*****No doctor registerd with the name of '+data[0]+'*****')
                print('*****please add the doctor first.*****')
                return False

    def __add_record(self,title,data):
        '''Add a record to a specific sheet.'''
        if(title=='Doctors' or title == 'Appointments'):
           if not self.__create_dependent_field_if_not_exists(title,data): return
        sheet = self.workbook[title]
        max_row = sheet.max_row
        max_id = sheet.cell(row=max_row,column=1).value
        if type(max_id) != int: max_id = 0
        for i in range(len(data)+1):
            sheet.cell(row=max_row+1,column=i+1).value = max_id+1 if i==0  else data[i-1]
        self.workbook.save('db.xlsx')

    def __prompt_addition(self,title):
        '''Prompt user input for adding a record in a specific sheet.'''
        if title == 'Patients':
            n = self.__validate('patient name',re.compile(r'^[a-zA-z]{1}.*'),'*****Invalid Name! name should start with character*****')
            ad = self.__validate('patient address',re.compile(r'^[a-zA-z]{1}.*'),'*****Invalid Address! address should start with character*****')
            a = self.__validate('patient age',re.compile(r'\d+'),'*****Invalid Age! age must be a number*****')
            self.__add_record(title,[n,ad,a])
        elif title == 'Doctors':
            n = self.__validate('doctor name',re.compile(r'^[a-zA-z]{1}.*'),'*****Invalid Name! name should start with character*****')
            d = self.__validate('department',re.compile(r'^[a-zA-z]{1}.*'),'*****Invalid Department! department should start with character*****')
            self.__add_record(title,[n,d])
        elif title == 'Appointments':
            dc = self.__validate('doctor name',re.compile(r'^[a-zA-z]{1}.*'),'*****Invalid Name! name should start with character*****')
            p = self.__validate('patient name',re.compile(r'^[a-zA-z]{1}.*'),'*****Invalid Name! name should start with character*****')
            d = date.today()
            self.__add_record(title,[dc,p,d])
        elif title == 'Departments':
            n = self.__validate('department name',re.compile(r'^[a-zA-z]{1}.*'),'*****Invalid Name! name should start with character*****')
            self.__add_record(title,[n])

    def __view(self,title):
        '''View all records of a specific sheet and justifies rows/columns for a better view.'''
        sheet = self.workbook[title]
        if sheet.max_row==1: print('\n*****No '+title+' registered yet!*****\n'); return
        items = []
        maxLens = []
        print('\n'+title+'\n')
        for i in range(1,sheet.max_column+1):
            rowLen = 0
            row = []
            for j in range(1,sheet.max_row+1):
                rowValue = sheet.cell(row=j,column=i).value
                if rowValue==None:break
                row.append(rowValue)
                if len(str(rowValue))> rowLen: rowLen=len(str(rowValue))
            else:
                items.append(row)
                maxLens.append(rowLen)
        for i in range(len(items[0])):
            x=''
            for j in range(len(items)):
                x += str(items[j][i]).ljust(maxLens[j]+10)
            print(x)

    def __delete_record(self,title,id):
        '''Deletes a record of a specific sheet by id and brings other rows on step up'''
        sheet = self.workbook[title]
        if sheet.max_row==1: print('\n*****No '+title+' registered yet!*****\n');return
        for i in range(2,sheet.max_row+1):
            if sheet.cell(row=i,column=1).value == id:
                if i < sheet.max_row:
                    for j in range(i+1,sheet.max_row+1):
                        diff = (j-i-1)+i
                        for k in range(1,sheet.max_column+1):
                            sheet.cell(row=diff,column=k).value = sheet.cell(row=j,column=k).value
                sheet.delete_rows(sheet.max_row,1)
                self.workbook.save('db.xlsx')
                break
        else: print('\n*****No '+title+' found with the id '+str(id)+'*****')


    def start(self):
        '''Program entry point.'''
        while True:
            print('\n\n\n',' Hospital Management System '.center(46,'*'))
            print('\n\n','Select your rule'.center(36,'-'))
            print('\n1: Admin\n2: User\n\nEnter Q to quit.\n')
            x = input('')
            if x == 'q' or x=='Q': print('\n*****bye*****.\n'); break
            elif x=='1':
                while True:
                    a = input('\nEnter Admin Password: ')
                    if self.__authenticate('B2',a): self.rule = 'admin'; self.__show_admin_menu(); break
                    else: print('\n*****Incorrect password!*****\n')
            elif x=='2':
                while True:
                    u = input('\nEnter User Password: ')
                    if self.__authenticate('B3',u): self.rule = 'user'; self.__show_admin_menu(); break
                    else: print('\n*****Incorrect password!*****\n')
            else: print('\n*****Invalid input!*****\n')

    
        
